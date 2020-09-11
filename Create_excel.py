import unittest2
from openpyxl import load_workbook , cell
import openpyxl
class Test(unittest2.TestCase):

    def test_read_excel_column_file(self):
        wb = openpyxl.Workbook()
        sheet = wb.create_sheet("DummySheet", 1)
        sheet["C1"].value = "This is sample writing"
        wb.save("C:/Desktop/Qlicr_Engine/Book19.xlsx")


if __name__ == "__main__":
    unittest2.main()